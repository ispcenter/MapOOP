'''
MapOOP v0.7 by Klykov Leonid
Development started 2020/06/18
Targets:
a) ± convenient and fast plotting of tkr characteristics	- надо сделать автонастраиваемые оси и выбор файлов не через код
b) + possibility of their combination
c) ± aproximation of experimental data - works bad: only for visualisation lines. NEED TO OPTIMIZE ALGORITHM!!! 27sec while 3 maps
d) - randomazer
e) - plotting KPDk levels
f) - livetime plotting by changing data -> in v2.0
g) + opening any excel files
h) + 2 different figures: turbine or Tk
'''

# =============================== 1) Head ============================

from matplotlib.ticker import (MultipleLocator, FormatStrFormatter, AutoMinorLocator)
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from csv import reader

def polynomialAproximation(x, y, X, dim):
	pc = np.polyfit(x, y, dim)

	if dim == 2:
		a = pc[0]
		b = pc[1]
		c = pc[2]
		Y = a*X**2 + b*X + c
		return Y

	elif dim == 3:
		a = pc[0]
		b = pc[1]
		c = pc[2]
		d = pc[3]
		Y = a*X**3 + b*X**2 + c*X + d
		return Y

	elif dim == 4:
		a = pc[0]
		b = pc[1]
		c = pc[2]
		d = pc[3]
		e = pc[4]
		Y = a*X**4 + b*X**3 + c*X**2 + d*X + e
		return Y

	elif dim == 5:
		a = pc[0]
		b = pc[1]
		c = pc[2]
		d = pc[3]
		e = pc[4]
		f = pc[5]
		Y = a*X**5 + b*X**4 + c*X**3 + d*X**2 + e*X + f
		return Y

	elif dim == 8:
		a = pc[0]
		b = pc[1]
		c = pc[2]
		d = pc[3]
		e = pc[4]
		f = pc[5]
		g = pc[6]
		h = pc[7]
		i = pc[8]
		Y = a*X**8 + b*X**7 + c*X**6 + d*X**5 + e*X**4 + f*X**3 + g*X**2 + h*X + i
		return Y

def PkAproximation(Gv, Pk):
	#1 ----------- initializing arrays ------------
	GvMaxT = np.linspace(1.005*Gv[0], 1.5*Gv[0], 1000)
	errors = []
	PC = []

	#2 ------------- searching Jmin ---------------
	for g in GvMaxT:
		pc = np.polyfit(Gv, Pk*(Gv - g), 3)
		a = pc[0]
		b = pc[1]
		c = pc[2]
		d = pc[3]
		x = Gv

		predictions = (a*x**3 + b*x**2 + c*x + d)/(x - g)
		J = sum((predictions - Pk)**2)/len(Gv)
		errors.append(J)
		PC.append((a, b, c, d))

	errors = np.array(errors)
	PC = np.array(PC)

	i = np.argmin(errors)
	a = PC[i][0]
	b = PC[i][1]
	c = PC[i][2]
	d = PC[i][3]

	#3 ---------- result --------------
	GvMaxT_opt = round(GvMaxT[i], 4)
	x = np.linspace(Gv[0], Gv[-1], count)
	y = (a*x**3 + b*x**2 + c*x + d)/(x - GvMaxT_opt)

	return np.array(y)

def opener(file):

	if file.endswith('xlsx'):
		wb = load_workbook(file)

	elif file.endswith('xls'):
		# читаем исходный файл xls и находим нужный лист
		xls_wb = open_workbook(file)
		xls_ws = xls_wb.sheet_by_index(0)

		# создаём openpyxl-приемник для данных из каждой ячейки
		wb = Workbook()
		ws = wb.active

		# заполняем его
		for row in range(xls_ws.nrows):
			new_row = []
			for cell in range(xls_ws.ncols):
				val = xls_ws.cell_value(row,cell)
				try:
					val = float(val)
				except ValueError:
					pass
				new_row.append(val)
			ws.append(new_row)

	else:
		print('Sorry, you had selected nonconformat type of file (it must be excel format like xlsx, xls or csv)')

	return wb

def setFigure_compressor_turbine():

	#1 ================= structure of a figure =================
	global fig, ax1, ax2, ax3, ax4, ax5, ax6
	fig = plt.figure(f'Характеристики {file_1}' + f', {file_2}')
	fig.subplots_adjust(left=0.04, bottom=0.06, right=0.99, top=0.97, wspace=0.09, hspace=0) # margins/paddings between subplots
	ax1 = plt.subplot2grid((3, 2), (0, 0), rowspan=2)
	ax2 = plt.subplot2grid((3, 2), (2, 0))
	ax3 = plt.subplot2grid((4, 2), (0, 1))
	ax4 = plt.subplot2grid((4, 2), (1, 1))
	ax5 = plt.subplot2grid((4, 2), (2, 1))
	ax6 = plt.subplot2grid((4, 2), (3, 1))

	#2 ===================== axes settings =====================

	#2.1 -------------------------Pk---------------------
	ax1.set_title('Характеристика компрессора')
	ax1.xaxis.set_major_locator(MultipleLocator(0.05))
	ax1.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax1.xaxis.set_minor_locator(MultipleLocator(0.01))
	ax1.set_xticklabels([])

	ax1.set_ylabel('Пк')
	ax1.yaxis.set_major_locator(MultipleLocator(0.4))
	ax1.yaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax1.yaxis.set_minor_locator(MultipleLocator(0.2))

	#2.2 -------------------------KPDk--------------------
	ax2.set_xlabel('Gв.пр')
	ax2.xaxis.set_major_locator(MultipleLocator(0.05))
	ax2.xaxis.set_major_formatter(FormatStrFormatter('%.2f'))
	ax2.xaxis.set_minor_locator(MultipleLocator(0.01))

	ax2.set_ylabel('КПД')
	ax2.yaxis.set_major_locator(MultipleLocator(0.1))
	ax2.yaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax2.yaxis.set_minor_locator(MultipleLocator(0.02))

	#2.3 --------------------------mft---------------------
	ax3.set_title('Характеристика турбины')
	ax3.xaxis.set_major_locator(MultipleLocator(0.2))
	ax3.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax3.xaxis.set_minor_locator(MultipleLocator(0.05))

	ax3.set_ylabel('mft')
	ax3.yaxis.set_major_locator(MultipleLocator(2))
	ax3.yaxis.set_major_formatter(FormatStrFormatter('%.0f'))
	ax3.yaxis.set_minor_locator(MultipleLocator(1))

	#2.4 --------------------------UCo---------------------
	ax4.xaxis.set_major_locator(MultipleLocator(0.2))
	ax4.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax4.xaxis.set_minor_locator(MultipleLocator(0.05))

	ax4.set_ylabel(r'$\frac{U_{т1}}{С_0}$')
	ax4.yaxis.set_major_locator(MultipleLocator(0.05))
	ax4.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
	ax4.yaxis.set_minor_locator(MultipleLocator(0.01))

	#2.5 ---------------------------Gr---------------------
	ax5.xaxis.set_major_locator(MultipleLocator(0.2))
	ax5.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax5.xaxis.set_minor_locator(MultipleLocator(0.05))

	ax5.set_ylabel(r'$\ G_{г.пр.}$')
	ax5.yaxis.set_major_locator(MultipleLocator(0.5))
	ax5.yaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax5.yaxis.set_minor_locator(MultipleLocator(0.1))

	#2.6 ---------------------------KPDt-------------------
	ax6.set_xlabel(r'$\pi_t$')
	ax6.xaxis.set_major_locator(MultipleLocator(0.2))
	ax6.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax6.xaxis.set_minor_locator(MultipleLocator(0.1))

	ax6.set_ylabel('KPDt')
	ax6.yaxis.set_major_locator(MultipleLocator(0.1))
	ax6.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
	ax6.yaxis.set_minor_locator(MultipleLocator(0.02))

	#3 ====================== grid settings =======================
	ax1.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax1.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
	ax2.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax2.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)

	ax3.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax3.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
	ax4.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax4.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
	ax5.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax5.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
	ax6.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax6.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)

def setFigure_Tk():
	#1 ================= structure of a figure =================
	global fig2, ax7, ax8, ax9, ax10
	fig2 = plt.figure('графики температур')
	fig2.subplots_adjust(left=0.04, bottom=0.06, right=0.99, top=0.97, wspace=0.09, hspace=0) # margins/paddings between subplots
	ax7 = plt.subplot2grid((3, 2), (0, 0), rowspan=3)
	ax8 = plt.subplot2grid((3, 2), (0, 1))
	ax9 = plt.subplot2grid((3, 2), (1, 1))
	ax10 = plt.subplot2grid((3, 2), (2, 1))

	#2 ===================== axes settings =====================

	#2.1 -------------------- Tk2 <- Tk1 --------------------
	ax7.set_title('Температуры')
	ax7.set_xlabel('Tk1')
	ax7.xaxis.set_major_locator(MultipleLocator(1))
	ax7.xaxis.set_major_formatter(FormatStrFormatter('%.0f'))
	ax7.xaxis.set_minor_locator(MultipleLocator(0.2))

	ax7.set_ylabel('Tk2')
	ax7.yaxis.set_major_locator(MultipleLocator(10))
	ax7.yaxis.set_major_formatter(FormatStrFormatter('%.0f'))
	ax7.yaxis.set_minor_locator(MultipleLocator(2))

	#2.2 ---------------------- Tk1 <- Gv -------------------
	ax8.set_title('Температурные зависимости')
	ax8.xaxis.set_major_locator(MultipleLocator(0.1))
	ax8.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax8.xaxis.set_minor_locator(MultipleLocator(0.02))
	ax8.set_xticklabels([])


	ax8.set_ylabel('Tk1')
	ax8.yaxis.set_major_locator(MultipleLocator(2))
	ax8.yaxis.set_major_formatter(FormatStrFormatter('%.0f'))
	ax8.yaxis.set_minor_locator(MultipleLocator(1))

	#2.3 ---------------------- Tk2 <- Gv --------------------
	ax9.xaxis.set_major_locator(MultipleLocator(0.1))
	ax9.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax9.xaxis.set_minor_locator(MultipleLocator(0.05))
	ax9.set_xticklabels([])

	ax9.set_ylabel('Tk2')
	ax9.yaxis.set_major_locator(MultipleLocator(50))
	ax9.yaxis.set_major_formatter(FormatStrFormatter('%.0f'))
	ax9.yaxis.set_minor_locator(MultipleLocator(10))

	#2.4 ----------------------- ΔTk <- Gv --------------------
	ax10.set_xlabel('Gv')
	ax10.xaxis.set_major_locator(MultipleLocator(0.1))
	ax10.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
	ax10.xaxis.set_minor_locator(MultipleLocator(0.05))

	ax10.set_ylabel('ΔTk')
	ax10.yaxis.set_major_locator(MultipleLocator(50))
	ax10.yaxis.set_major_formatter(FormatStrFormatter('%.0f'))
	ax10.yaxis.set_minor_locator(MultipleLocator(10))

	#3 ====================== grid settings =======================
	ax7.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax7.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
	ax8.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax8.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)

	ax9.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax9.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
	ax10.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
	ax10.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)

class Plotter:
	"""plotter grafics. Помпаж не передал ещё"""
	def __init__(self,
			Gv, Pk, KPDk, Tk1, Tk2, ΔTk, 			# original compressor data
			GvX, PkY, KPDkY, Tk1X, Tk2Y, ΔTkY,	# aproximated data
			Tk1Z, Tk2Z, ΔTkZ,							# tk <- Gv grafs
			Pt, mft, UCo, Gr, KPDt,					# original turbine data
			PtX, mftY, UCoY, GrY, KPDtY,			# aproximated data
			marker, color):							# settings

		#1 get data
		self.Gv = Gv
		self.Pk = Pk
		self.KPDk = KPDk
		self.Tk1 = Tk1
		self.Tk2 = Tk2
		self.ΔTk = ΔTk

		self.GvX = GvX
		self.PkY = PkY
		self.KPDkY = KPDkY
		self.Tk1X = Tk1X
		self.Tk2Y = Tk2Y
		self.ΔTkY = ΔTkY

		self.Tk1Z = Tk1Z
		self.Tk2Z = Tk2Z
		self.ΔTkZ = ΔTkZ

		self.Pt = Pt
		self.mft = mft
		self.UCo = UCo
		self.Gr = Gr
		self.KPDt = KPDt

		self.PtX = PtX
		self.mftY = mftY
		self.UCoY = UCoY
		self.GrY = GrY
		self.KPDtY = KPDtY

		self.marker = marker
		self.color = color

		#2 plot compressor
		self.ax1plot = ax1.plot(self.Gv, self.Pk,   self.marker, self.GvX, self.PkY,   '-', c=self.color, markersize = markersize)
		self.ax2plot = ax2.plot(self.Gv, self.KPDk, self.marker, self.GvX, self.KPDkY, '-', c=self.color, markersize = markersize)

		#3 plot turbine:
		self.ax3plot = ax3.plot(self.Pt, self.mft, self.marker, self.PtX, self.mftY, '-', c=self.color, markersize = markersize)
		self.ax4plot = ax4.plot(self.Pt, self.UCo, self.marker, self.PtX, self.UCoY, '-', c=self.color, markersize = markersize)
		self.ax5plot = ax5.plot(self.Pt, self.Gr,  self.marker, self.PtX, self.GrY, '-', c=self.color, markersize = markersize)
		self.ax6plot = ax6.plot(self.Pt, self.KPDt,self.marker, self.PtX, self.KPDtY, '-', c=self.color, markersize = markersize)

		#4 plot Tk
		if show_Tk != 0:
			self.ax7plot = ax7.plot(self.Tk1, self.Tk2, self.marker, self.Tk1X, self.Tk2Y, '-', c=self.color, markersize = markersize)
			self.ax8plot = ax8.plot(self.Gv, self.Tk1, self.marker, self.GvX, self.Tk1Z, '-', c=self.color, markersize = markersize)
			self.ax9plot = ax9.plot(self.Gv, self.Tk2, self.marker, self.GvX, self.Tk2Z, '-', c=self.color, markersize = markersize)
			self.ax10plot = ax10.plot(self.Gv, self.ΔTk, self.marker, self.GvX, self.ΔTkZ, '-', c=self.color, markersize = markersize)

class Uk2:
	'''Prototype of a vetka
	1) Receive dataset
	2) Activates all data grafs'''

	def __init__(self, ws, uLabel, marker, color, firstRow, lastRow):
		#1 ------------------ Data preparing --------------------
		self.ws = ws
		self.uLabel = uLabel
		self.marker = marker
		self.color = color
		self.firstRow = firstRow
		self.lastRow = lastRow
		self.data = self.ws[f'A{self.firstRow}':f'AR{self.lastRow}']
		self.data = [row for row in self.data if (type(row[1].value) == float and row[1].value != 0)] #if not check => promlems with 0 and str values

		#2 ----------------- Compressor data --------------------
		self.Gv   = np.array([row[1].value for row in self.data])
		self.Pk   = np.array([row[2].value for row in self.data])
		self.KPDk = np.array([row[3].value for row in self.data])
		self.Tk1  = np.array([row[41].value for row in self.data])
		self.Tk2  = np.array([row[42].value for row in self.data])
		self.ΔTk  = np.array([row[43].value for row in self.data])

		self.GvX = np.linspace(self.Gv[0], self.Gv[-1], num=count, endpoint=True)
		self.PkY = PkAproximation(self.Gv, self.Pk)

		self.Tk1X = np.linspace(min(self.Tk1), max(self.Tk1), count)
		self.Tk2Y = polynomialAproximation(self.Tk1, self.Tk2, self.Tk1X, 2)
		self.ΔTkY = self.Tk2Y - self.Tk1X

		self.Tk1Z = polynomialAproximation(self.Gv, self.Tk1, self.GvX, 4)	#(x, y, X, dim)
		self.Tk2Z = polynomialAproximation(self.Gv, self.Tk2, self.GvX, 4)
		self.ΔTkZ = polynomialAproximation(self.Gv, self.ΔTk, self.GvX, 4)

		self.KPDkY = self.Tk1Z*(self.PkY**(0.4/1.4)-1)/self.ΔTkZ

		#3 -------------------- Turbine data ---------------------
		self.Pt = np.array([(row[4].value) for row in self.data if (type(row[4].value) == float and row[4].value != 0)])
		self.PtX = np.linspace(min(self.Pt), max(self.Pt), num=count, endpoint=True)

		self.KPDt  = np.array([row[6].value for row in self.data if (type(row[6].value) == float and row[6].value != 0)])
		self.Gr  = np.array([row[5].value for row in self.data if (type(row[5].value) == float and row[5].value != 0)])
		self.UCo  = np.array([row[7].value for row in self.data if (type(row[7].value) == float and row[7].value != 0)])
		self.mft  = np.array([row[8].value for row in self.data if (type(row[8].value) == float and row[8].value != 0)])

		self.KPDtY= polynomialAproximation(self.Pt, self.KPDt,self.PtX, 2)	#(x, y, X, dim)
		self.GrY  = polynomialAproximation(self.Pt, self.Gr,  self.PtX, 2)
		self.UCoY = polynomialAproximation(self.Pt, self.UCo, self.PtX, 2)
		self.mftY = polynomialAproximation(self.Pt, self.mft, self.PtX, 2)

		#4 ------------------- Plotting results ---------------------
		self.plotter = Plotter(
			self.Gv, self.Pk, self.KPDk, self.Tk1, self.Tk2, self.ΔTk,
			self.GvX, self.PkY, self.KPDkY, self.Tk1X, self.Tk2Y, self.ΔTkY,
			self.Tk1Z, self.Tk2Z, self.ΔTkZ,
			self.Pt, self.mft, self.UCo, self.Gr, self.KPDt,
			self.PtX, self.mftY, self.UCoY, self.GrY, self.KPDtY,
			self.marker, self.color)

class Map:
	"""Prototype of a mapData
	1) Receive file and color
	2) Activates all uk2
	3) Activates pompaz
	4) KPDk levels"""

	def __init__(self, file, color):
		#1 color setup
		self.color = color

		#2 data setup
		self.file = file
		self.wb = opener(file)
		self.ws = self.wb.active
		self.uList = []															# obj base - only not empty uk2 will go here

		#3 all uk2 activation
		if (type(self.ws[40][1].value) == float and self.ws[40][1].value != 0):	# do Uk2 creation or not, if it is empty
			self.u200 = Uk2(self.ws, 200, '1', self.color, 40, 57)					# (ws, label, marker, color, firstRow, lastRow)
			self.uList.append(self.u200)														# if uk2 obj was created, we say to object base about it
		if (type(self.ws[59][1].value) == float and self.ws[59][1].value != 0):
			self.u250 = Uk2(self.ws, 250, '^', self.color, 59, 76)
			self.uList.append(self.u250)
		if (type(self.ws[78][1].value) == float and self.ws[78][1].value != 0):
			self.u300 = Uk2(self.ws, 300, 's', self.color, 78, 95)
			self.uList.append(self.u300)
		if (type(self.ws[97][1].value) == float and self.ws[97][1].value != 0):
			self.u350 = Uk2(self.ws, 350, 'D', self.color, 97, 114)
			self.uList.append(self.u350)
		if (type(self.ws[116][1].value) == float and self.ws[116][1].value != 0):
			self.u400 = Uk2(self.ws, 400, 'x', self.color, 116, 133)
			self.uList.append(self.u400)
		if (type(self.ws[135][1].value) == float and self.ws[135][1].value != 0):
			self.u450 = Uk2(self.ws, 450, 'o', self.color, 135, 152)
			self.uList.append(self.u450)
		if (type(self.ws[154][1].value) == float and self.ws[154][1].value != 0):
			self.u500 = Uk2(self.ws, 500, '+', self.color, 154, 171)
			self.uList.append(self.u500)
		if (type(self.ws[173][1].value) == float and self.ws[173][1].value != 0):
			self.u550 = Uk2(self.ws, 550, 'v', self.color, 173, 190)
			self.uList.append(self.u550)
		if (type(self.ws[192][1].value) == float and self.ws[192][1].value != 0):
			self.u600 = Uk2(self.ws, 600, '2', self.color, 192, 209)
			self.uList.append(self.u600)

		#4 get pompaz coordinates
		self.xpomp = [u.Gv[-1] for u in self.uList]
		self.ypomp = [u.Pk.data[-1] for u in self.uList]

		#5 compute KPDk max line
		self.x_KPDk_Max = tuple([u.GvX[np.argmax(u.KPDk)] for u in self.uList])		#tuple n x 1
		self.y_KPDk_Max = tuple([u.Pk[np.argmax(u.KPDk)] for u in self.uList])		#tuple n x 1
		self.z_KPDk_Max = tuple([u.KPDk[np.argmax(u.KPDk)] for u in self.uList])	#tuple n x 1

		#6 plot pompaz
		ax1.plot(self.xpomp, self.ypomp, c=self.color)

		#7 compute levels
		levels = (58, 60, 64, 68, 70, 72, 74, 76, 77, 78, 79, 80)

		# for level in levels:
		# 	if map1.KPDk.any()

mapsList = []

# =============================== 2) Config ===============================
file_1 = 'ТКРы/60к6.xls'
file_2 = 'ТКРы/60к7.xls'
file_3 = ''
file_4 = ''

color_1 = 'black'	# color of map1
color_2 = 'red'	# color of map2
color_3 = 'blue'	# color of map3
color_4 = 'green'	# color of map4

show_Tk = 0
markersize = 4		# size of markers in graphs
count = 111			# number of aproximation points

# ============================ 3) Create the map ============================
setFigure_compressor_turbine()
if show_Tk != 0: setFigure_Tk()

if len(file_1) > 0: map1 = Map(file_1, color_1)
if len(file_2) > 0: map2 = Map(file_2, color_2)
if len(file_3) > 0: map3 = Map(file_3, color_3)
if len(file_4) > 0: map4 = Map(file_4, color_4)

plt.show()
