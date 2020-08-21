'''
MapOOP v0.3 by Klykov Leonid
Development started 2020/06/18
Targets:
a) ± convenient and fast plotting of tkr characteristics	- надо сделать автонастраиваемые оси и выбор файлов не через код
b) + possibility of their combination
c) ± aproximation of experimental data - badly? but works
d) - randomazer
e) - plotting KPDk levels
f) - livetime d) by changing data
'''

# --------------------------Head------------------------

from matplotlib.ticker import (MultipleLocator, FormatStrFormatter, AutoMinorLocator)
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from time import sleep

class Compressor:
	'''Prototype of a compressor graf
	1) Receive dataset
	2) Get data from it
	3) Minimize equation using data'''

	def __init__(self, dataset, column, uLabel, Gv, GvX):
		self.dataset = dataset
		self.column = column
		self.uLabel = uLabel
		self.Gv = Gv
		self.GvX = GvX

		self.uc200 = 1.5			#при d=5 оптимально 1.5
		self.uc600 = 1.001		#при d=5 оптимально 1.001-1,0001
		self.uc = np.polyfit([200,600],[self.uc200, self.uc600], 1)
		self.Gvmaxt = (self.uc[0]*uLabel + self.uc[1])*self.Gv[0] # теоретический максимум по расходу. это не очень точно, но хотя бы работает. В будущем надо обязательно попробовать градиентный спуск
		self.data  = np.array([(row[column].value) for row in self.dataset if row[column].value != None])
		
		self.pc = np.polyfit(self.Gv, self.data*(self.Gv - self.Gvmaxt), 3)
		self.a = self.pc[0]
		self.b = self.pc[1]
		self.c = self.pc[2]
		self.d = self.pc[3]
		# self.e = self.pc[4]
		# self.f = self.pc[5]
		
		self.y = (self.a*self.GvX**3 + self.b*self.GvX**2 + self.c*self.GvX + self.d)/(self.GvX - self.Gvmaxt) #test with d=3 => underfitting
		# self.y = (self.a*self.GvX**4 + self.b*self.GvX**3 + self.c*self.GvX**2 + self.d*self.GvX + self.e)/(self.GvX - self.Gvmaxt)
		# self.y = (self.a*self.GvX**5 + self.b*self.GvX**4 + self.c*self.GvX**3 + self.d*self.GvX**2 + self.e*self.GvX + self.f)/(self.GvX - self.Gvmaxt)
		
	def aproximation5(self):
		print(self.data)

class Turbine:
	'''Prototype of a compressor graf
	1) Receive dataset
	2) Get data from it
	3) Minimize equation using data'''

	def __init__(self, dataset, column, Pt, PtX):
		self.dataset = dataset
		self.column = column
		self.Pt = Pt
		self.PtX = PtX
		self.data  = np.array([(row[column].value) for row in self.dataset if row[column].value != None])

		self.pc = np.polyfit(self.Pt, self.data, 2)
		self.a = self.pc[0]
		self.b = self.pc[1]
		self.c = self.pc[2]

		self.y = self.a*self.PtX**2 + self.b*self.PtX + self.c

	def aproximation2(self):
		print(self.data)

class Uk2:
	'''Prototype of a vetka
	1) Receive dataset
	2) Activates all data grafs'''

	def __init__(self, ws, uLabel, marker, firstRow, lastRow):

		self.ws = ws
		self.uLabel = uLabel
		self.marker = marker
		self.firstRow = firstRow
		self.lastRow = lastRow
		self.data = self.ws[f'A{self.firstRow}':f'I{self.lastRow}']

		self.Gv = np.array([(row[1].value) for row in self.data  if row[1].value != None]) #this check is necessary, because not all data rows may filled
		self.GvX = np.linspace(self.Gv[0], self.Gv[-1], num=count, endpoint=True)

		self.Pt = np.array([(row[4].value) for row in self.data  if row[4].value != None]) #this check is necessary, because not all data rows may filled
		self.PtX = np.linspace(min(self.Pt), max(self.Pt), num=count, endpoint=True)

		self.Pk 	 = Compressor(self.data, 2, self.uLabel, self.Gv, self.GvX) 	#(dataset, column, Gv, x_coords)
		self.KPDk = Compressor(self.data, 3, self.uLabel, self.Gv, self.GvX)	#(dataset, column, Gv, x_coords)

		self.KPDt = Turbine(self.data, 6, self.Pt, self.PtX)
		self.Gr 	 = Turbine(self.data, 5, self.Pt, self.PtX)
		self.UCo  = Turbine(self.data, 7, self.Pt, self.PtX)
		self.mft  = Turbine(self.data, 8, self.Pt, self.PtX)

class Map:
	"""Prototype of a mapData
	1) Receive file and color
	2) Activates all uk2
	3) Activates pompaz
	4) KPDk levels"""

	def __init__(self, file, color):
		#1 color setup
		self.color = color

		#2 data setup
		self.file = file
		self.wb = load_workbook(filename=file)
		self.ws = self.wb.active
		self.uList = []										# only not empty uk2 will go here

		#3 all uk2 activation
		if self.ws[40][1].value != None:					# very useful check - do Uk2 creation or not, if it empty
			self.u200 = Uk2(self.ws, 200, '1', 40, 57)# (ws, label, marker, firstRow, lastRow)
			self.uList.append(self.u200)					# if it created, we say it to object base
		if self.ws[59][1].value != None:
			self.u250 = Uk2(self.ws, 250, '^', 59, 76)
			self.uList.append(self.u250)
		if self.ws[78][1].value != None:
			self.u300 = Uk2(self.ws, 300, 's', 78, 95)
			self.uList.append(self.u300)
		if self.ws[97][1].value != None:
			self.u350 = Uk2(self.ws, 350, 'D', 97, 114)
			self.uList.append(self.u350)
		if self.ws[116][1].value != None:
			self.u400 = Uk2(self.ws, 400, 'x', 116, 133)
			self.uList.append(self.u400)
		if self.ws[135][1].value != None:
			self.u450 = Uk2(self.ws, 450, 'o', 135, 152)
			self.uList.append(self.u450)
		if self.ws[154][1].value != None:
			self.u500 = Uk2(self.ws, 500, '+', 154, 171)
			self.uList.append(self.u500)
		if self.ws[173][1].value != None:
			self.u550 = Uk2(self.ws, 550, 'v', 173, 190)
			self.uList.append(self.u550)
		if self.ws[192][1].value != None:
			self.u600 = Uk2(self.ws, 600, '2', 192, 209)
			self.uList.append(self.u600)

		#4 get pompaz coordinates
		self.xpomp = [u.Gv[-1] for u in self.uList]
		self.ypomp = [u.Pk.data[-1] for u in self.uList]

		#5 compute KPDk max line
		self.x_KPDk_Max = tuple([u.GvX[np.argmax(u.KPDk.y)] for u in self.uList])		#tuple n x 1
		self.y_KPDk_Max = tuple([u.Pk.y[np.argmax(u.KPDk.y)] for u in self.uList])		#tuple n x 1
		self.z_KPDk_Max = tuple([u.KPDk.y[np.argmax(u.KPDk.y)] for u in self.uList])	#tuple n x 1

		#6 compute levels
		levels = (58, 60, 64, 68, 70, 72, 74, 76, 77, 78, 79, 80)

		# for level in levels:
		# 	if map1.KPDk.any()

# --------------------------Config--------------------------
file_1 = 'ТКР_80.15.13к04_2020.06.04.xlsx'
file_2 = 'ТКР_80.15.13к05_2020.06.04.xlsx'
file_3 = 'file3'
file_4 = 'file4'

count = 100			# number of aproximation points
markersize = 4		# size of marker on grafs

color_1 = 'black'	# color of map1
color_2 = 'red'	# color of map2
color_3 = 'blue'	# color of map3
color_4 = 'green'	# color of map4

# ----------------------Create the map-----------------------
map1 = Map(file_1, color_1)
map2 = Map(file_2, color_2)
print(
	map1.u550.KPDk.a,
	map1.u550.KPDk.b,
	map1.u550.KPDk.c,
	map1.u550.KPDk.d,
	map1.u550.KPDk.Gvmaxt
	)
# ----------------------- Plot the map-----------------------

# structure of a figure
fig = plt.figure('Зависимость Тк от Gв.пр')
ax1 = plt.subplot2grid((3, 2), (0, 0), rowspan=2)
ax2 = plt.subplot2grid((3, 2), (2, 0))
ax3 = plt.subplot2grid((4, 2), (0, 1))
ax4 = plt.subplot2grid((4, 2), (1, 1))
ax5 = plt.subplot2grid((4, 2), (2, 1))
ax6 = plt.subplot2grid((4, 2), (3, 1))

# axes settings - Pk
ax1.xaxis.set_major_locator(MultipleLocator(0.1))
ax1.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax1.xaxis.set_minor_locator(MultipleLocator(0.02))
ax1.set_xticklabels([])

ax1.yaxis.set_major_locator(MultipleLocator(0.4))
ax1.yaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax1.yaxis.set_minor_locator(MultipleLocator(0.2))

# axes settings - KPDk
ax2.xaxis.set_major_locator(MultipleLocator(0.1))
ax2.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax2.xaxis.set_minor_locator(MultipleLocator(0.02))

ax2.yaxis.set_major_locator(MultipleLocator(0.1))
ax2.yaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax2.yaxis.set_minor_locator(MultipleLocator(0.02))

# axes settings - mft
ax3.xaxis.set_major_locator(MultipleLocator(0.2))
ax3.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax3.xaxis.set_minor_locator(MultipleLocator(0.05))

ax3.yaxis.set_major_locator(MultipleLocator(2))
ax3.yaxis.set_major_formatter(FormatStrFormatter('%.0f'))
ax3.yaxis.set_minor_locator(MultipleLocator(1))

# axes settings - UCo
ax4.xaxis.set_major_locator(MultipleLocator(0.2))
ax4.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax4.xaxis.set_minor_locator(MultipleLocator(0.05))

ax4.yaxis.set_major_locator(MultipleLocator(0.05))
ax4.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
ax4.yaxis.set_minor_locator(MultipleLocator(0.01))

# axes settings - Gr
ax5.xaxis.set_major_locator(MultipleLocator(0.2))
ax5.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax5.xaxis.set_minor_locator(MultipleLocator(0.05))

ax5.yaxis.set_major_locator(MultipleLocator(0.5))
ax5.yaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax5.yaxis.set_minor_locator(MultipleLocator(0.1))

# axes settings - KPDt
ax6.xaxis.set_major_locator(MultipleLocator(0.2))
ax6.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))
ax6.xaxis.set_minor_locator(MultipleLocator(0.1))

ax6.yaxis.set_major_locator(MultipleLocator(0.1))
ax6.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
ax6.yaxis.set_minor_locator(MultipleLocator(0.02))

# grid settings
ax1.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
ax1.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
ax2.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
ax2.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)

ax3.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
ax3.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
ax4.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
ax4.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
ax5.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
ax5.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)
ax6.grid(b=True, which='major', axis='both', color='#000000', linestyle='dotted', linewidth=1)
ax6.grid(b=True, which='minor', axis='both', color='#BDBDBD', linestyle='dotted', linewidth=1)

# margins/paddings
fig.subplots_adjust(left=0.04, bottom=0.06, right=0.99, top=0.97, wspace=0.09, hspace=0)

# plots
ax1.set_title('Характеристика компрессора')
ax1.set_ylabel('Пк')
ax1.plot(
	# map1.u200.Gv, map1.u200.Pk.data, map1.u200.marker, map1.u200.GvX, map1.u200.Pk.y, '-',
	# map1.u250.Gv, map1.u250.Pk.data, map1.u250.marker, map1.u250.GvX, map1.u250.Pk.y, '-',
	# map1.u300.Gv, map1.u300.Pk.data, map1.u300.marker, map1.u300.GvX, map1.u300.Pk.y, '-',
	# map1.u350.Gv, map1.u350.Pk.data, map1.u350.marker, map1.u350.GvX, map1.u350.Pk.y, '-',
	# map1.u400.Gv, map1.u400.Pk.data, map1.u400.marker, map1.u400.GvX, map1.u400.Pk.y, '-',
	# map1.u450.Gv, map1.u450.Pk.data, map1.u450.marker, map1.u450.GvX, map1.u450.Pk.y, '-',
	# map1.u500.Gv, map1.u500.Pk.data, map1.u500.marker, map1.u500.GvX, map1.u500.Pk.y, '-',
	# map1.u550.Gv, map1.u550.Pk.data, map1.u550.marker, map1.u550.GvX, map1.u550.Pk.y, '-',
	# map1.u600.Gv, map1.u600.Pk.data, map1.u600.marker, map1.u600.GvX, map1.u600.Pk.y, '-',
	map1.xpomp, map1.ypomp, '-', map1.x_KPDk_Max, map1.y_KPDk_Max, c=map1.color, markersize = markersize)
ax1.plot(
	# map2.u200.Gv, map2.u200.Pk.data, map2.u200.marker, map2.u200.GvX, map2.u200.Pk.y, '-',
	map2.u250.Gv, map2.u250.Pk.data, map2.u250.marker, map2.u250.GvX, map2.u250.Pk.y, '-',
	map2.u300.Gv, map2.u300.Pk.data, map2.u300.marker, map2.u300.GvX, map2.u300.Pk.y, '-',
	map2.u350.Gv, map2.u350.Pk.data, map2.u350.marker, map2.u350.GvX, map2.u350.Pk.y, '-',
	map2.u400.Gv, map2.u400.Pk.data, map2.u400.marker, map2.u400.GvX, map2.u400.Pk.y, '-',
	map2.u450.Gv, map2.u450.Pk.data, map2.u450.marker, map2.u450.GvX, map2.u450.Pk.y, '-',
	map2.u500.Gv, map2.u500.Pk.data, map2.u500.marker, map2.u500.GvX, map2.u500.Pk.y, '-',
	map2.u550.Gv, map2.u550.Pk.data, map2.u550.marker, map2.u550.GvX, map2.u550.Pk.y, '-',
	# map2.u600.Gv, map2.u600.Pk.data, map2.u600.marker, map2.u600.GvX, map2.u600.Pk.y, '-',
	map2.xpomp, map2.ypomp, map2.x_KPDk_Max, map2.y_KPDk_Max, c=map2.color, markersize = markersize)

ax2.set_xlabel('Gв.пр')
ax2.set_ylabel('КПД')
ax2.plot(
	# map1.u200.Gv, map1.u200.KPDk.data, map1.u200.marker, map1.u200.GvX, map1.u200.KPDk.y, '-',
	map1.u250.Gv, map1.u250.KPDk.data, map1.u250.marker, map1.u250.GvX, map1.u250.KPDk.y, '-',
	map1.u300.Gv, map1.u300.KPDk.data, map1.u300.marker, map1.u300.GvX, map1.u300.KPDk.y, '-',
	map1.u350.Gv, map1.u350.KPDk.data, map1.u350.marker, map1.u350.GvX, map1.u350.KPDk.y, '-',
	map1.u400.Gv, map1.u400.KPDk.data, map1.u400.marker, map1.u400.GvX, map1.u400.KPDk.y, '-',
	map1.u450.Gv, map1.u450.KPDk.data, map1.u450.marker, map1.u450.GvX, map1.u450.KPDk.y, '-',
	map1.u500.Gv, map1.u500.KPDk.data, map1.u500.marker, map1.u500.GvX, map1.u500.KPDk.y, '-',
	map1.u550.Gv, map1.u550.KPDk.data, map1.u550.marker, map1.u550.GvX, map1.u550.KPDk.y, '-',
	# map1.u600.Gv, map1.u600.KPDk.data, map1.u600.marker, map1.u600.GvX, map1.u600.KPDk.y, '-',
	c=map1.color, markersize = markersize)
ax2.plot(
	# map2.u200.Gv, map2.u200.KPDk.data, map2.u200.marker, map2.u200.GvX, map2.u200.KPDk.y, '-',
	map2.u250.Gv, map2.u250.KPDk.data, map2.u250.marker, map2.u250.GvX, map2.u250.KPDk.y, '-',
	map2.u300.Gv, map2.u300.KPDk.data, map2.u300.marker, map2.u300.GvX, map2.u300.KPDk.y, '-',
	map2.u350.Gv, map2.u350.KPDk.data, map2.u350.marker, map2.u350.GvX, map2.u350.KPDk.y, '-',
	map2.u400.Gv, map2.u400.KPDk.data, map2.u400.marker, map2.u400.GvX, map2.u400.KPDk.y, '-',
	map2.u450.Gv, map2.u450.KPDk.data, map2.u450.marker, map2.u450.GvX, map2.u450.KPDk.y, '-',
	map2.u500.Gv, map2.u500.KPDk.data, map2.u500.marker, map2.u500.GvX, map2.u500.KPDk.y, '-',
	map2.u550.Gv, map2.u550.KPDk.data, map2.u550.marker, map2.u550.GvX, map2.u550.KPDk.y, '-',
	# map2.u600.Gv, map2.u600.KPDk.data, map2.u600.marker, map2.u600.GvX, map2.u600.KPDk.y, '-',
	c=map2.color, markersize = markersize)

ax3.set_title('Характеристика турбины')
ax3.set_ylabel('mft')
ax3.plot(
	# map1.u200.Pt, map1.u200.mft.data, map1.u200.marker, map1.u200.PtX, map1.u200.mft.y, '-',
	map1.u250.Pt, map1.u250.mft.data, map1.u250.marker, map1.u250.PtX, map1.u250.mft.y, '-',
	map1.u300.Pt, map1.u300.mft.data, map1.u300.marker, map1.u300.PtX, map1.u300.mft.y, '-',
	map1.u350.Pt, map1.u350.mft.data, map1.u350.marker, map1.u350.PtX, map1.u350.mft.y, '-',
	map1.u400.Pt, map1.u400.mft.data, map1.u400.marker, map1.u400.PtX, map1.u400.mft.y, '-',
	map1.u450.Pt, map1.u450.mft.data, map1.u450.marker, map1.u450.PtX, map1.u450.mft.y, '-',
	map1.u500.Pt, map1.u500.mft.data, map1.u500.marker, map1.u500.PtX, map1.u500.mft.y, '-',
	map1.u550.Pt, map1.u550.mft.data, map1.u550.marker, map1.u550.PtX, map1.u550.mft.y, '-',
	# map1.u600.Pt, map1.u600.mft.data, map1.u600.marker, map1.u600.PtX, map1.u600.mft.y, '-',
	c=map1.color, markersize = markersize)

ax4.set_ylabel(r'$\frac{U_{т1}}{С_0}$')
ax4.plot(
	# map1.u200.Pt, map1.u200.UCo.data, map1.u200.marker, map1.u200.PtX, map1.u200.UCo.y, '-',
	map1.u250.Pt, map1.u250.UCo.data, map1.u250.marker, map1.u250.PtX, map1.u250.UCo.y, '-',
	map1.u300.Pt, map1.u300.UCo.data, map1.u300.marker, map1.u300.PtX, map1.u300.UCo.y, '-',
	map1.u350.Pt, map1.u350.UCo.data, map1.u350.marker, map1.u350.PtX, map1.u350.UCo.y, '-',
	map1.u400.Pt, map1.u400.UCo.data, map1.u400.marker, map1.u400.PtX, map1.u400.UCo.y, '-',
	map1.u450.Pt, map1.u450.UCo.data, map1.u450.marker, map1.u450.PtX, map1.u450.UCo.y, '-',
	map1.u500.Pt, map1.u500.UCo.data, map1.u500.marker, map1.u500.PtX, map1.u500.UCo.y, '-',
	map1.u550.Pt, map1.u550.UCo.data, map1.u550.marker, map1.u550.PtX, map1.u550.UCo.y, '-',
	# map1.u600.Pt, map1.u600.UCo.data, map1.u600.marker, map1.u600.PtX, map1.u600.UCo.y, '-',
	c=map1.color, markersize = markersize)

ax5.set_ylabel(r'$\ G_{г.пр.}$')
ax5.plot(
	# map1.u200.Pt, map1.u200.Gr.data, map1.u200.marker, map1.u200.PtX, map1.u200.Gr.y, '-',
	map1.u250.Pt, map1.u250.Gr.data, map1.u250.marker, map1.u250.PtX, map1.u250.Gr.y, '-',
	map1.u300.Pt, map1.u300.Gr.data, map1.u300.marker, map1.u300.PtX, map1.u300.Gr.y, '-',
	map1.u350.Pt, map1.u350.Gr.data, map1.u350.marker, map1.u350.PtX, map1.u350.Gr.y, '-',
	map1.u400.Pt, map1.u400.Gr.data, map1.u400.marker, map1.u400.PtX, map1.u400.Gr.y, '-',
	map1.u450.Pt, map1.u450.Gr.data, map1.u450.marker, map1.u450.PtX, map1.u450.Gr.y, '-',
	map1.u500.Pt, map1.u500.Gr.data, map1.u500.marker, map1.u500.PtX, map1.u500.Gr.y, '-',
	map1.u550.Pt, map1.u550.Gr.data, map1.u550.marker, map1.u550.PtX, map1.u550.Gr.y, '-',
	# map1.u600.Pt, map1.u600.Gr.data, map1.u600.marker, map1.u600.PtX, map1.u600.Gr.y, '-',
	c=map1.color, markersize = markersize)

ax6.set_xlabel(r'$\pi_t$')
ax6.set_ylabel('KPDt')
ax6.plot(
	# map1.u200.Pt, map1.u200.KPDt.data, map1.u200.marker, map1.u200.PtX, map1.u200.KPDt.y, '-',
	map1.u250.Pt, map1.u250.KPDt.data, map1.u250.marker, map1.u250.PtX, map1.u250.KPDt.y, '-',
	map1.u300.Pt, map1.u300.KPDt.data, map1.u300.marker, map1.u300.PtX, map1.u300.KPDt.y, '-',
	map1.u350.Pt, map1.u350.KPDt.data, map1.u350.marker, map1.u350.PtX, map1.u350.KPDt.y, '-',
	map1.u400.Pt, map1.u400.KPDt.data, map1.u400.marker, map1.u400.PtX, map1.u400.KPDt.y, '-',
	map1.u450.Pt, map1.u450.KPDt.data, map1.u450.marker, map1.u450.PtX, map1.u450.KPDt.y, '-',
	map1.u500.Pt, map1.u500.KPDt.data, map1.u500.marker, map1.u500.PtX, map1.u500.KPDt.y, '-',
	map1.u550.Pt, map1.u550.KPDt.data, map1.u550.marker, map1.u550.PtX, map1.u550.KPDt.y, '-',
	# map1.u600.Pt, map1.u600.KPDt.data, map1.u600.marker, map1.u600.PtX, map1.u600.KPDt.y, '-',
	c=map1.color, markersize = markersize)

plt.show()
# ---------------------------testRESULT--------------------------
# print(map1.u250.Pk.data)
